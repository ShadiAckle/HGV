#!/usr/bin/env python3
"""Generate HGV Compensation Hub data model Excel workbook from DDL definitions."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs" / "HGV_Comp_Data_Model.xlsx"

CATALOG = "edw_dev_hris"
SCHEMA = "hgv_comp"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def hdr(ws, row: int, cols: list[str]) -> None:
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = max(width, 10)


# (entity, object_type, layer, storage_prod, grain, description, ddl_source)
ENTITIES = [
    ("_src_tour_spine", "VIEW", "Staging", "VIEW", "tour_key_hash", "Filtered Cognos tour detail spine (36mo lookback in script 15)", "12_bootstrap_live_source_views.sql"),
    ("_src_rep_directory", "VIEW", "Staging", "VIEW", "rep_id", "Union marketing + field rep keys from Cognos/PwC", "12_bootstrap_live_source_views.sql"),
    ("_src_period_calendar", "VIEW", "Staging", "VIEW", "period_id", "Quarter calendar from commissions + tours", "12_bootstrap_live_source_views.sql"),
    ("dim_team", "TABLE/VIEW", "Dimension", "VIEW (prod)", "team_id", "Sales team", "01_create_schema.sql / 12 views"),
    ("dim_rep", "TABLE/VIEW", "Dimension", "TABLE (script 16)", "rep_id", "All reps: marketing + field (FY2025–26 when materialized)", "01 + 16_materialize_marketing_core.sql"),
    ("dim_marketing_rep", "TABLE", "Dimension", "TABLE (script 16)", "rep_id", "Marketing channel rep picker (FY2025–26 tours)", "16_materialize_marketing_core.sql"),
    ("dim_period", "TABLE/VIEW", "Dimension", "TABLE (script 16)", "period_id", "Pay/reporting quarters (FY2025–26 when materialized)", "01 + 16_materialize_marketing_core.sql"),
    ("dim_plan_version", "TABLE/VIEW", "Dimension", "VIEW (prod)", "plan_version_id", "Compensation plan version", "01_create_schema.sql"),
    ("dim_product_line", "TABLE/VIEW", "Dimension", "VIEW (prod)", "product_line_id", "Product line (FFS flag)", "01_create_schema.sql"),
    ("dim_finance_period", "TABLE", "Dimension", "DELTA", "period_id", "Finance budget, accrual, ROI thresholds", "05b_extend_finance_reference.sql"),
    ("dim_location", "TABLE/VIEW", "Dimension", "VIEW (prod)", "location_id", "Properties, sales centers, desks", "09_create_guest_registry.sql"),
    ("dim_household", "TABLE/VIEW", "Dimension", "VIEW (prod)", "household_id", "Household demographics (banded)", "09_create_guest_registry.sql"),
    ("dim_guest", "TABLE/VIEW", "Dimension", "VIEW (prod)", "guest_id", "Guest spine for tour enrichment", "09_create_guest_registry.sql"),
    ("bridge_tour_guest", "TABLE/VIEW", "Bridge", "VIEW (prod)", "tour_id + guest_id", "Tour-to-guest M:N bridge", "09_create_guest_registry.sql"),
    ("fact_quota_attainment", "TABLE/VIEW", "Fact", "VIEW (prod)", "rep_id + period_id", "Rep quota and attainment", "01_create_schema.sql"),
    ("fact_payout", "TABLE/VIEW", "Fact", "VIEW (prod)", "rep_id + period_id", "Rep payout breakdown", "01_create_schema.sql"),
    ("fact_deal_credit", "TABLE/VIEW", "Fact", "VIEW (prod)", "deal_id", "Deal-level quota credit", "01_create_schema.sql"),
    ("fact_team_snapshot", "TABLE/VIEW", "Fact", "VIEW (prod)", "team_id + period_id", "Manager team KPI snapshot", "01_create_schema.sql"),
    ("fact_rep_product_mix", "TABLE/VIEW", "Fact", "VIEW (prod)", "rep_id + period_id + product_line_id", "Rep product mix %", "01_create_schema.sql"),
    ("fact_plan_eligibility", "TABLE/VIEW", "Fact", "VIEW (prod)", "rep_id + period_id", "Plan assignment and proration", "05_extend_admin_finance.sql"),
    ("fact_comp_admin_log", "TABLE", "Fact", "DELTA", "event_id", "Comp admin audit log (writable)", "05_extend_admin_finance.sql"),
    ("fact_chargeback", "TABLE/VIEW", "Fact", "VIEW (prod)", "chargeback_id", "Field sales chargebacks per deal", "05_extend_admin_finance.sql"),
    ("fact_tour_quality", "TABLE/VIEW", "Fact", "VIEW (prod)", "tour_id", "Tour quality for finance analysis", "05_extend_admin_finance.sql"),
    ("fact_marketing_rep_period", "TABLE", "Fact", "TABLE (script 16)", "rep_id + period_id", "Marketing rep period KPI rollup", "06 + 16_materialize_marketing_core.sql"),
    ("fact_marketing_rep_metric", "TABLE/VIEW", "Fact", "VIEW on Delta", "rep_id + period_id + metric_name", "Marketing plan metric weights", "06 + 16 views"),
    ("fact_marketing_tour_payout", "TABLE", "Fact", "TABLE (script 16)", "tour_id", "Marketing tour ledger and payouts", "06 + 16_materialize_marketing_core.sql"),
    ("fact_marketing_chargeback", "TABLE/VIEW", "Fact", "VIEW on Delta", "chargeback_id", "Marketing tour chargebacks", "06 + 16 views"),
    ("fact_marketing_arrival", "TABLE/VIEW", "Fact", "VIEW on Delta", "arrival_id", "Upcoming arrivals / projected pay", "06 + 16 views"),
    ("fact_rep_market_position", "TABLE/VIEW", "Fact", "VIEW (prod)", "rep_id + period_id", "Rep pay vs industry benchmark", "06_create_marketing_benchmark.sql"),
    ("fact_regional_bonus_area", "TABLE", "Fact", "DELTA", "area_id + period_id", "Regional bonus area volume vs budget", "07_create_regional_bonus.sql"),
    ("fact_regional_bonus_tier", "TABLE", "Fact", "DELTA", "area_id + period_id + level", "Regional bonus tier breakdown", "07_create_regional_bonus.sql"),
    ("fact_guest_ownership", "TABLE/VIEW", "Fact", "VIEW (prod)", "ownership_id", "Guest HGV ownership interests", "09_create_guest_registry.sql"),
    ("fact_guest_rental_stay", "TABLE/VIEW", "Fact", "VIEW (prod)", "stay_id", "Guest rental/owner stays", "09_create_guest_registry.sql"),
    ("fact_guest_tour_history", "TABLE/VIEW", "Fact", "VIEW (prod)", "history_id", "Prior tour outcomes per guest", "09_create_guest_registry.sql"),
    ("industry_comp_benchmark", "TABLE", "Reference", "DELTA (seed)", "benchmark_id", "Industry comp benchmarks", "06_create_marketing_benchmark.sql"),
    ("plan_assessment_profile", "TABLE", "Reference", "DELTA (seed)", "persona_id + plan_id", "Plan assessment header", "10_create_plan_assessment.sql"),
    ("plan_assessment_segment", "TABLE", "Reference", "DELTA (seed)", "persona_id + attribute + segment_order", "HGV vs market plan segments", "10_create_plan_assessment.sql"),
    ("scenario_run", "TABLE", "Scenario", "DELTA (writable)", "scenario_id", "What-if scenario inputs", "01_create_schema.sql"),
    ("scenario_result", "TABLE", "Scenario", "DELTA (writable)", "scenario_id", "What-if scenario outputs", "01_create_schema.sql"),
    ("scenario_payout_series", "TABLE", "Scenario", "DELTA (writable)", "scenario_id + bucket_order", "Scenario chart series", "01_create_schema.sql"),
]

# (entity, column, data_type, nullable, is_pk, description)
COLUMNS: list[tuple[str, str, str, str, str, str]] = [
    # dim_team
    ("dim_team", "team_id", "STRING", "N", "Y", "Team surrogate key"),
    ("dim_team", "team_name", "STRING", "N", "N", "Team display name"),
    ("dim_team", "region", "STRING", "N", "N", "Geographic region"),
    # dim_rep
    ("dim_rep", "rep_id", "STRING", "N", "Y", "Rep / participant ID"),
    ("dim_rep", "rep_name", "STRING", "N", "N", "Rep display name"),
    ("dim_rep", "level_code", "STRING", "N", "N", "Job level (L5–L9, MKT)"),
    ("dim_rep", "team_id", "STRING", "N", "N", "FK → dim_team.team_id"),
    ("dim_rep", "manager_rep_id", "STRING", "Y", "N", "FK → dim_rep.rep_id (manager)"),
    ("dim_rep", "region", "STRING", "N", "N", "Assigned region"),
    ("dim_rep", "is_active", "BOOLEAN", "N", "N", "Active on roster"),
    # dim_marketing_rep
    ("dim_marketing_rep", "rep_id", "STRING", "N", "Y", "Marketing rep employee ID"),
    ("dim_marketing_rep", "rep_name", "STRING", "N", "N", "Rep name"),
    ("dim_marketing_rep", "level_code", "STRING", "N", "N", "Always MKT for marketing channel"),
    ("dim_marketing_rep", "team_id", "STRING", "N", "N", "Sales team code"),
    ("dim_marketing_rep", "region", "STRING", "N", "N", "Office region"),
    ("dim_marketing_rep", "is_active", "BOOLEAN", "N", "N", "Active flag"),
    # dim_period
    ("dim_period", "period_id", "STRING", "N", "Y", "Quarter key e.g. 2026-Q2"),
    ("dim_period", "period_label", "STRING", "N", "N", "Display label e.g. Q2 2026"),
    ("dim_period", "period_start", "DATE", "N", "N", "Quarter start date"),
    ("dim_period", "period_end", "DATE", "N", "N", "Quarter end date"),
    ("dim_period", "is_current", "BOOLEAN", "N", "N", "Current period flag"),
    # dim_plan_version
    ("dim_plan_version", "plan_version_id", "STRING", "N", "Y", "Plan version key"),
    ("dim_plan_version", "plan_name", "STRING", "N", "N", "Plan name"),
    ("dim_plan_version", "effective_start", "DATE", "N", "N", "Plan effective start"),
    ("dim_plan_version", "effective_end", "DATE", "Y", "N", "Plan effective end"),
    # dim_product_line
    ("dim_product_line", "product_line_id", "STRING", "N", "Y", "Product line key"),
    ("dim_product_line", "product_line_name", "STRING", "N", "N", "Product line name"),
    ("dim_product_line", "is_ffs", "BOOLEAN", "N", "N", "Fee-for-service flag"),
    # dim_finance_period
    ("dim_finance_period", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("dim_finance_period", "budget_comp", "DECIMAL(14,2)", "N", "N", "Comp budget for period"),
    ("dim_finance_period", "payroll_lock_date", "DATE", "N", "N", "Payroll lock date"),
    ("dim_finance_period", "accrual_basis", "STRING", "N", "N", "Accrual basis code"),
    ("dim_finance_period", "var_comp_target_min_pct", "DECIMAL(5,2)", "N", "N", "Variable comp corridor min %"),
    ("dim_finance_period", "var_comp_target_max_pct", "DECIMAL(5,2)", "N", "N", "Variable comp corridor max %"),
    ("dim_finance_period", "spiff_roi_threshold", "DECIMAL(5,2)", "N", "N", "SPIFF ROI threshold"),
    ("dim_finance_period", "rescission_target_pct", "DECIMAL(5,2)", "N", "N", "Rescission target %"),
    ("dim_finance_period", "ebitda_margin_pct", "DECIMAL(5,2)", "N", "N", "EBITDA margin %"),
    ("dim_finance_period", "ffs_reserve_pct", "DECIMAL(5,2)", "N", "N", "FFS reserve %"),
    ("dim_finance_period", "accrual_policy_notes", "STRING", "Y", "N", "Finance policy notes"),
    # fact_quota_attainment
    ("fact_quota_attainment", "rep_id", "STRING", "N", "Y", "FK → dim_rep.rep_id"),
    ("fact_quota_attainment", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_quota_attainment", "plan_version_id", "STRING", "N", "N", "FK → dim_plan_version"),
    ("fact_quota_attainment", "quota_amount", "DECIMAL(14,2)", "N", "N", "Quota $"),
    ("fact_quota_attainment", "credited_amount", "DECIMAL(14,2)", "N", "N", "Credited $"),
    ("fact_quota_attainment", "attainment_pct", "DECIMAL(6,2)", "N", "N", "Attainment %"),
    ("fact_quota_attainment", "deals_closed_count", "INT", "N", "N", "Closed deal count"),
    ("fact_quota_attainment", "next_tier_threshold_pct", "DECIMAL(6,2)", "N", "N", "Next tier threshold %"),
    ("fact_quota_attainment", "next_tier_gap_amount", "DECIMAL(14,2)", "N", "N", "Gap to next tier $"),
    # fact_payout
    ("fact_payout", "rep_id", "STRING", "N", "Y", "FK → dim_rep.rep_id"),
    ("fact_payout", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_payout", "base_pay", "DECIMAL(14,2)", "N", "N", "Base pay"),
    ("fact_payout", "commission", "DECIMAL(14,2)", "N", "N", "Commission"),
    ("fact_payout", "bonus", "DECIMAL(14,2)", "N", "N", "Bonus"),
    ("fact_payout", "total_earnings", "DECIMAL(14,2)", "N", "N", "Total earnings"),
    ("fact_payout", "total_paid", "DECIMAL(14,2)", "N", "N", "Total paid"),
    # fact_deal_credit
    ("fact_deal_credit", "deal_id", "STRING", "N", "Y", "Deal / contract key"),
    ("fact_deal_credit", "rep_id", "STRING", "N", "N", "FK → dim_rep.rep_id"),
    ("fact_deal_credit", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_deal_credit", "product_line_id", "STRING", "N", "N", "FK → dim_product_line"),
    ("fact_deal_credit", "property_code", "STRING", "N", "N", "Property code"),
    ("fact_deal_credit", "property_display_name", "STRING", "N", "N", "Property name"),
    ("fact_deal_credit", "credit_amount", "DECIMAL(14,2)", "N", "N", "Credit amount"),
    ("fact_deal_credit", "credit_status", "STRING", "N", "N", "Credit status"),
    ("fact_deal_credit", "credit_date", "DATE", "N", "N", "Credit date"),
    # fact_team_snapshot
    ("fact_team_snapshot", "team_id", "STRING", "N", "Y", "FK → dim_team.team_id"),
    ("fact_team_snapshot", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_team_snapshot", "team_attainment_pct", "DECIMAL(6,2)", "N", "N", "Team attainment %"),
    ("fact_team_snapshot", "top_performer_count", "INT", "N", "N", "Top performer count"),
    ("fact_team_snapshot", "at_risk_count", "INT", "N", "N", "At-risk rep count"),
    ("fact_team_snapshot", "ffs_sales_pct", "DECIMAL(6,2)", "N", "N", "FFS sales %"),
    ("fact_team_snapshot", "ffs_target_pct", "DECIMAL(6,2)", "N", "N", "FFS target %"),
    # fact_marketing_rep_period
    ("fact_marketing_rep_period", "rep_id", "STRING", "N", "Y", "FK → dim_marketing_rep.rep_id"),
    ("fact_marketing_rep_period", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_marketing_rep_period", "rep_name", "STRING", "Y", "N", "Rep name denorm"),
    ("fact_marketing_rep_period", "plan_id", "STRING", "Y", "N", "Marketing plan ID"),
    ("fact_marketing_rep_period", "assigned_area", "STRING", "Y", "N", "Assigned bonus area"),
    ("fact_marketing_rep_period", "bonus_area_id", "STRING", "Y", "N", "Regional bonus area"),
    ("fact_marketing_rep_period", "qtd_earnings", "DECIMAL(14,2)", "Y", "N", "QTD earnings"),
    ("fact_marketing_rep_period", "paid_to_date", "DECIMAL(14,2)", "Y", "N", "Paid to date"),
    ("fact_marketing_rep_period", "qualified_tours", "INT", "Y", "N", "Qualified tour count"),
    ("fact_marketing_rep_period", "tours_shown", "INT", "Y", "N", "Tours shown count"),
    ("fact_marketing_rep_period", "show_rate_pct", "DECIMAL(6,2)", "Y", "N", "Show rate %"),
    ("fact_marketing_rep_period", "penetration_pct", "DECIMAL(6,2)", "Y", "N", "Penetration %"),
    ("fact_marketing_rep_period", "penetration_target_pct", "DECIMAL(6,2)", "Y", "N", "Penetration target %"),
    ("fact_marketing_rep_period", "spiff_active", "BOOLEAN", "Y", "N", "SPIFF tier active"),
    ("fact_marketing_rep_period", "next_tier_label", "STRING", "Y", "N", "Next tier label"),
    ("fact_marketing_rep_period", "next_tier_gap_tours", "INT", "Y", "N", "Tours to next tier"),
    ("fact_marketing_rep_period", "qualified_tour_pay", "DECIMAL(14,2)", "Y", "N", "Qualified tour pay"),
    ("fact_marketing_rep_period", "courtesy_tour_pay", "DECIMAL(14,2)", "Y", "N", "Courtesy tour pay"),
    ("fact_marketing_rep_period", "penetration_spiff", "DECIMAL(14,2)", "Y", "N", "Penetration SPIFF"),
    ("fact_marketing_rep_period", "chargebacks", "DECIMAL(14,2)", "Y", "N", "Chargeback total"),
    ("fact_marketing_rep_period", "total_payout", "DECIMAL(14,2)", "Y", "N", "Total payout"),
    ("fact_marketing_rep_period", "base_pct", "DECIMAL(6,2)", "Y", "N", "Base % of TCC"),
    ("fact_marketing_rep_period", "variable_pct", "DECIMAL(6,2)", "Y", "N", "Variable % of TCC"),
    ("fact_marketing_rep_period", "tcc_gap_vs_market_pct", "DECIMAL(6,2)", "Y", "N", "TCC gap vs market"),
    # fact_marketing_tour_payout
    ("fact_marketing_tour_payout", "tour_id", "STRING", "N", "Y", "Tour ID"),
    ("fact_marketing_tour_payout", "rep_id", "STRING", "N", "N", "FK → dim_marketing_rep.rep_id"),
    ("fact_marketing_tour_payout", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_marketing_tour_payout", "guest_name", "STRING", "Y", "N", "Guest display name"),
    ("fact_marketing_tour_payout", "guest_type", "STRING", "Y", "N", "Qualified / Showed / Courtesy"),
    ("fact_marketing_tour_payout", "arrival_date", "DATE", "Y", "N", "Tour / arrival date"),
    ("fact_marketing_tour_payout", "tour_status", "STRING", "Y", "N", "Tour status"),
    ("fact_marketing_tour_payout", "code", "STRING", "Y", "N", "Channel / desk code"),
    ("fact_marketing_tour_payout", "payout", "DECIMAL(14,2)", "Y", "N", "Tour payout $"),
    ("fact_marketing_tour_payout", "fps_eligible", "BOOLEAN", "Y", "N", "FPS eligible flag"),
    ("fact_marketing_tour_payout", "fps_potential", "DECIMAL(14,2)", "Y", "N", "FPS potential $"),
    ("fact_marketing_tour_payout", "notes", "STRING", "Y", "N", "Marketing program notes"),
    ("fact_marketing_tour_payout", "guest_id", "STRING", "Y", "N", "FK → dim_guest.guest_id"),
    ("fact_marketing_tour_payout", "household_id", "STRING", "Y", "N", "FK → dim_household"),
    ("fact_marketing_tour_payout", "planned_tour_location_id", "STRING", "Y", "N", "FK → dim_location"),
    ("fact_marketing_tour_payout", "current_stay_location_id", "STRING", "Y", "N", "Current stay location"),
    ("fact_marketing_tour_payout", "lead_source", "STRING", "Y", "N", "Lead source"),
    ("fact_marketing_tour_payout", "abc_score", "STRING", "Y", "N", "FICO tier / ABC score"),
    ("fact_marketing_tour_payout", "package_type", "STRING", "Y", "N", "Marketing package type"),
    ("fact_marketing_tour_payout", "xref_tour_id", "STRING", "Y", "N", "Cross-ref tour key"),
    ("fact_marketing_tour_payout", "tour_booked_date", "DATE", "Y", "N", "Booked date"),
    ("fact_marketing_tour_payout", "rep_name", "STRING", "Y", "N", "Rep name (materialized denorm)"),
    ("fact_marketing_tour_payout", "rep_region", "STRING", "Y", "N", "Rep region (materialized)"),
    ("fact_marketing_tour_payout", "rep_team_id", "STRING", "Y", "N", "Rep team (materialized)"),
    # scenario_run
    ("scenario_run", "scenario_id", "STRING", "N", "Y", "Scenario key"),
    ("scenario_run", "scenario_name", "STRING", "N", "N", "Scenario name"),
    ("scenario_run", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("scenario_run", "quota_change_pct", "DECIMAL(6,2)", "N", "N", "Quota change lever %"),
    ("scenario_run", "commission_rate_pct", "DECIMAL(6,2)", "N", "N", "Commission rate %"),
    ("scenario_run", "bonus_rate_change_pct", "DECIMAL(6,2)", "N", "N", "Bonus rate change %"),
    ("scenario_run", "accelerator_change_pct", "DECIMAL(6,2)", "N", "N", "Accelerator change %"),
    ("scenario_run", "tour_volume_change_pct", "DECIMAL(6,2)", "N", "N", "Tour volume lever %"),
    ("scenario_run", "conversion_rate_change_pct", "DECIMAL(6,2)", "N", "N", "Conversion lever %"),
    ("scenario_run", "created_by", "STRING", "N", "N", "Creator user ID"),
    # scenario_result
    ("scenario_result", "scenario_id", "STRING", "N", "Y", "FK → scenario_run.scenario_id"),
    ("scenario_result", "projected_payouts", "DECIMAL(16,2)", "N", "N", "Projected payouts"),
    ("scenario_result", "budget_impact", "DECIMAL(16,2)", "N", "N", "Budget impact"),
    ("scenario_result", "projected_cost", "DECIMAL(16,2)", "N", "N", "Projected cost"),
    ("scenario_result", "expected_performance_pct", "DECIMAL(6,2)", "N", "N", "Expected performance %"),
    # fact_comp_admin_log
    ("fact_comp_admin_log", "event_id", "STRING", "N", "Y", "Event key"),
    ("fact_comp_admin_log", "rep_id", "STRING", "N", "N", "FK → dim_rep.rep_id"),
    ("fact_comp_admin_log", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_comp_admin_log", "event_type", "STRING", "N", "N", "Event type code"),
    ("fact_comp_admin_log", "amount", "DECIMAL(14,2)", "Y", "N", "Adjustment amount"),
    ("fact_comp_admin_log", "reason", "STRING", "N", "N", "Reason text"),
    ("fact_comp_admin_log", "approved_by", "STRING", "Y", "N", "Approver"),
    ("fact_comp_admin_log", "created_at", "TIMESTAMP", "N", "N", "Event timestamp"),
    ("fact_comp_admin_log", "attributed_nsv", "DECIMAL(14,2)", "Y", "N", "Attributed NSV for SPIFF"),
    # fact_marketing_rep_metric
    ("fact_marketing_rep_metric", "rep_id", "STRING", "N", "Y", "FK → dim_marketing_rep.rep_id"),
    ("fact_marketing_rep_metric", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_marketing_rep_metric", "metric_name", "STRING", "N", "Y", "Plan metric name"),
    ("fact_marketing_rep_metric", "weight_pct", "DECIMAL(6,2)", "Y", "N", "Metric weight %"),
    ("fact_marketing_rep_metric", "earnings", "DECIMAL(14,2)", "Y", "N", "Earnings attributed to metric"),
    ("fact_marketing_rep_metric", "attainment_pct", "DECIMAL(6,2)", "Y", "N", "Attainment %"),
    ("fact_marketing_rep_metric", "target_label", "STRING", "Y", "N", "Target description"),
    ("fact_marketing_rep_metric", "opportunity_usd", "DECIMAL(14,2)", "Y", "N", "Upside opportunity $"),
    # fact_marketing_chargeback
    ("fact_marketing_chargeback", "chargeback_id", "STRING", "N", "Y", "Chargeback key"),
    ("fact_marketing_chargeback", "rep_id", "STRING", "N", "N", "FK → dim_marketing_rep.rep_id"),
    ("fact_marketing_chargeback", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_marketing_chargeback", "guest_name", "STRING", "Y", "N", "Guest name"),
    ("fact_marketing_chargeback", "tour_id", "STRING", "Y", "N", "FK → fact_marketing_tour_payout.tour_id"),
    ("fact_marketing_chargeback", "premium_gift", "STRING", "Y", "N", "Premium gift reference"),
    ("fact_marketing_chargeback", "chargeback_amount", "DECIMAL(14,2)", "Y", "N", "Chargeback $"),
    ("fact_marketing_chargeback", "notes", "STRING", "Y", "N", "Notes"),
    # fact_marketing_arrival
    ("fact_marketing_arrival", "arrival_id", "STRING", "N", "Y", "Arrival key"),
    ("fact_marketing_arrival", "rep_id", "STRING", "N", "N", "FK → dim_marketing_rep.rep_id"),
    ("fact_marketing_arrival", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_marketing_arrival", "guest_name", "STRING", "Y", "N", "Guest name"),
    ("fact_marketing_arrival", "guest_type", "STRING", "Y", "N", "Guest type"),
    ("fact_marketing_arrival", "arrival_datetime", "STRING", "Y", "N", "Arrival datetime"),
    ("fact_marketing_arrival", "desk", "STRING", "Y", "N", "Desk assignment"),
    ("fact_marketing_arrival", "potential_qualified_tour", "DECIMAL(14,2)", "Y", "N", "Potential qualified tour $"),
    ("fact_marketing_arrival", "potential_fps_payout", "DECIMAL(14,2)", "Y", "N", "Potential FPS $"),
    ("fact_marketing_arrival", "projected_total_payout", "DECIMAL(14,2)", "Y", "N", "Projected total $"),
    # fact_rep_market_position
    ("fact_rep_market_position", "rep_id", "STRING", "N", "Y", "FK → dim_rep.rep_id"),
    ("fact_rep_market_position", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_rep_market_position", "rep_name", "STRING", "Y", "N", "Rep name"),
    ("fact_rep_market_position", "role_key", "STRING", "Y", "N", "Role benchmark key"),
    ("fact_rep_market_position", "tcc_gap_vs_market_pct", "DECIMAL(6,2)", "Y", "N", "TCC gap vs market %"),
    ("fact_rep_market_position", "base_pct", "DECIMAL(6,2)", "Y", "N", "Base %"),
    ("fact_rep_market_position", "variable_pct", "DECIMAL(6,2)", "Y", "N", "Variable %"),
    ("fact_rep_market_position", "quota_attainment_pct", "DECIMAL(6,2)", "Y", "N", "Quota attainment %"),
    # regional bonus
    ("fact_regional_bonus_area", "area_id", "STRING", "N", "Y", "Regional bonus area ID"),
    ("fact_regional_bonus_area", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_regional_bonus_area", "site_line", "STRING", "Y", "N", "Site line label"),
    ("fact_regional_bonus_area", "smt_volume", "DECIMAL(16,2)", "Y", "N", "SMT volume"),
    ("fact_regional_bonus_area", "budget_volume", "DECIMAL(16,2)", "Y", "N", "Budget volume"),
    ("fact_regional_bonus_area", "volume_var_pct", "DECIMAL(6,2)", "Y", "N", "Volume variance %"),
    ("fact_regional_bonus_tier", "area_id", "STRING", "N", "Y", "FK → fact_regional_bonus_area.area_id"),
    ("fact_regional_bonus_tier", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_regional_bonus_tier", "level", "INT", "N", "Y", "Bonus tier level"),
    ("fact_regional_bonus_tier", "salespeople_count", "INT", "Y", "N", "Headcount in tier"),
    ("fact_regional_bonus_tier", "avg_tier_volume", "DECIMAL(16,2)", "Y", "N", "Avg tier volume"),
    ("fact_regional_bonus_tier", "total_tier_volume", "DECIMAL(16,2)", "Y", "N", "Total tier volume"),
    ("fact_regional_bonus_tier", "total_cmi", "DECIMAL(16,2)", "Y", "N", "Total CMI"),
    ("fact_regional_bonus_tier", "cost_pct", "DECIMAL(6,2)", "Y", "N", "Cost %"),
    # guest registry
    ("dim_guest", "guest_id", "STRING", "N", "Y", "Guest surrogate key"),
    ("dim_guest", "guest_name", "STRING", "N", "N", "Guest display name"),
    ("dim_guest", "email", "STRING", "Y", "N", "Email (tokenized in prod)"),
    ("dim_guest", "phone_token", "STRING", "Y", "N", "Phone token"),
    ("dim_guest", "guest_type", "STRING", "N", "N", "Guest type"),
    ("dim_guest", "owner_flag", "BOOLEAN", "N", "N", "HGV owner flag"),
    ("dim_guest", "household_id", "STRING", "Y", "N", "FK → dim_household.household_id"),
    ("dim_guest", "qualification_code", "STRING", "Y", "N", "Qualification code"),
    ("dim_guest", "tour_booked_date", "DATE", "Y", "N", "Tour booked date"),
    ("dim_household", "household_id", "STRING", "N", "Y", "Household key"),
    ("dim_household", "hh_size_band", "STRING", "N", "N", "Household size band"),
    ("dim_household", "income_band", "STRING", "N", "N", "Income band"),
    ("dim_household", "home_msa", "STRING", "Y", "N", "Home MSA"),
    ("dim_household", "enrichment_source", "STRING", "Y", "N", "Enrichment vendor"),
    ("dim_household", "enrichment_as_of", "DATE", "Y", "N", "Enrichment as-of date"),
    ("dim_location", "location_id", "STRING", "N", "Y", "Location key"),
    ("dim_location", "location_name", "STRING", "N", "N", "Location name"),
    ("dim_location", "location_type", "STRING", "N", "N", "Property / sales center / desk"),
    ("dim_location", "market", "STRING", "N", "N", "Market"),
    ("dim_location", "brand", "STRING", "N", "N", "Brand"),
    ("dim_location", "desk_label", "STRING", "Y", "N", "Desk label"),
    ("bridge_tour_guest", "tour_id", "STRING", "N", "Y", "FK → tour"),
    ("bridge_tour_guest", "guest_id", "STRING", "N", "Y", "FK → dim_guest.guest_id"),
    ("bridge_tour_guest", "is_primary", "BOOLEAN", "N", "N", "Primary guest on tour"),
    # plan assessment + benchmarks
    ("industry_comp_benchmark", "benchmark_id", "STRING", "N", "Y", "Benchmark row key"),
    ("industry_comp_benchmark", "role_key", "STRING", "N", "N", "Role key"),
    ("industry_comp_benchmark", "role_label", "STRING", "Y", "N", "Role label"),
    ("industry_comp_benchmark", "metric_code", "STRING", "N", "N", "Metric code"),
    ("industry_comp_benchmark", "market_value", "DECIMAL(10,2)", "Y", "N", "Market benchmark value"),
    ("industry_comp_benchmark", "hgv_typical_value", "DECIMAL(10,2)", "Y", "N", "HGV typical value"),
    ("industry_comp_benchmark", "unit", "STRING", "Y", "N", "Unit of measure"),
    ("industry_comp_benchmark", "benchmark_source", "STRING", "Y", "N", "Source citation"),
    ("industry_comp_benchmark", "effective_period", "STRING", "Y", "N", "Effective period"),
    ("industry_comp_benchmark", "notes", "STRING", "Y", "N", "Notes"),
    ("plan_assessment_profile", "persona_id", "STRING", "N", "Y", "Persona key"),
    ("plan_assessment_profile", "plan_id", "STRING", "N", "Y", "Plan key"),
    ("plan_assessment_profile", "role_title", "STRING", "N", "N", "Role title"),
    ("plan_assessment_profile", "channel_code", "STRING", "N", "N", "Channel code"),
    ("plan_assessment_profile", "effective_period", "STRING", "N", "N", "Effective period"),
    ("plan_assessment_segment", "persona_id", "STRING", "N", "Y", "FK → plan_assessment_profile"),
    ("plan_assessment_segment", "effective_period", "STRING", "N", "Y", "Effective period"),
    ("plan_assessment_segment", "attribute", "STRING", "N", "Y", "Attribute name"),
    ("plan_assessment_segment", "attribute_order", "INT", "N", "N", "Attribute sort order"),
    ("plan_assessment_segment", "side", "STRING", "N", "N", "HGV vs market side"),
    ("plan_assessment_segment", "segment_order", "INT", "N", "Y", "Segment order"),
    ("plan_assessment_segment", "segment_label", "STRING", "Y", "N", "Segment label"),
    ("plan_assessment_segment", "segment_value", "STRING", "N", "N", "Segment value"),
    # admin / finance facts
    ("fact_plan_eligibility", "rep_id", "STRING", "N", "Y", "FK → dim_rep.rep_id"),
    ("fact_plan_eligibility", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_plan_eligibility", "plan_version_id", "STRING", "N", "N", "FK → dim_plan_version"),
    ("fact_plan_eligibility", "job_code", "STRING", "N", "N", "Job code"),
    ("fact_plan_eligibility", "location_code", "STRING", "N", "N", "Location code"),
    ("fact_plan_eligibility", "brand", "STRING", "N", "N", "Brand"),
    ("fact_plan_eligibility", "effective_start", "DATE", "N", "N", "Eligibility start"),
    ("fact_plan_eligibility", "effective_end", "DATE", "Y", "N", "Eligibility end"),
    ("fact_plan_eligibility", "proration_pct", "DECIMAL(5,2)", "N", "N", "Proration %"),
    ("fact_plan_eligibility", "eligibility_flag", "BOOLEAN", "N", "N", "Eligible flag"),
    ("fact_plan_eligibility", "exclusion_reason", "STRING", "Y", "N", "Exclusion reason"),
    ("fact_chargeback", "chargeback_id", "STRING", "N", "Y", "Chargeback key"),
    ("fact_chargeback", "deal_id", "STRING", "N", "N", "Deal ID"),
    ("fact_chargeback", "rep_id", "STRING", "N", "N", "FK → dim_rep.rep_id"),
    ("fact_chargeback", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_chargeback", "original_commission", "DECIMAL(14,2)", "N", "N", "Original commission"),
    ("fact_chargeback", "chargeback_amount", "DECIMAL(14,2)", "N", "N", "Chargeback amount"),
    ("fact_chargeback", "reserve_held", "DECIMAL(14,2)", "N", "N", "Reserve held"),
    ("fact_chargeback", "reserve_released", "DECIMAL(14,2)", "N", "N", "Reserve released"),
    ("fact_chargeback", "reason", "STRING", "N", "N", "Reason"),
    ("fact_chargeback", "status", "STRING", "N", "N", "Status"),
    ("fact_tour_quality", "tour_id", "STRING", "N", "Y", "Tour ID"),
    ("fact_tour_quality", "rep_id", "STRING", "N", "N", "FK → dim_rep.rep_id"),
    ("fact_tour_quality", "period_id", "STRING", "N", "N", "FK → dim_period.period_id"),
    ("fact_tour_quality", "lead_source", "STRING", "N", "N", "Lead source"),
    ("fact_tour_quality", "abc_score", "STRING", "N", "N", "ABC / FICO tier"),
    ("fact_tour_quality", "package_type", "STRING", "N", "N", "Package type"),
    ("fact_tour_quality", "showed_flag", "BOOLEAN", "N", "N", "Showed flag"),
    ("fact_tour_quality", "closed_flag", "BOOLEAN", "N", "N", "Closed flag"),
    ("fact_tour_quality", "contract_status", "STRING", "N", "N", "Contract status"),
    ("fact_tour_quality", "rescission_flag", "BOOLEAN", "N", "N", "Rescission flag"),
    ("fact_tour_quality", "net_sales_volume", "DECIMAL(14,2)", "N", "N", "Net sales volume"),
    ("fact_tour_quality", "vpg", "DECIMAL(10,2)", "N", "N", "Volume per guest"),
    ("fact_tour_quality", "ebitda_estimate", "DECIMAL(14,2)", "N", "N", "EBITDA estimate"),
    ("fact_rep_product_mix", "rep_id", "STRING", "N", "Y", "FK → dim_rep.rep_id"),
    ("fact_rep_product_mix", "period_id", "STRING", "N", "Y", "FK → dim_period.period_id"),
    ("fact_rep_product_mix", "product_line_id", "STRING", "N", "Y", "FK → dim_product_line"),
    ("fact_rep_product_mix", "mix_pct", "DECIMAL(6,2)", "N", "N", "Mix %"),
    ("scenario_payout_series", "scenario_id", "STRING", "N", "Y", "FK → scenario_run.scenario_id"),
    ("scenario_payout_series", "series_label", "STRING", "N", "N", "Series label"),
    ("scenario_payout_series", "bucket_order", "INT", "N", "Y", "Chart bucket order"),
    ("scenario_payout_series", "bucket_label", "STRING", "N", "N", "Bucket label"),
    ("scenario_payout_series", "payout_amount", "DECIMAL(16,2)", "N", "N", "Payout amount"),
]

RELATIONSHIPS = [
    ("dim_rep", "team_id", "dim_team", "team_id", "many-to-one"),
    ("dim_rep", "manager_rep_id", "dim_rep", "rep_id", "many-to-one"),
    ("fact_quota_attainment", "rep_id", "dim_rep", "rep_id", "many-to-one"),
    ("fact_quota_attainment", "period_id", "dim_period", "period_id", "many-to-one"),
    ("fact_payout", "rep_id", "dim_rep", "rep_id", "many-to-one"),
    ("fact_payout", "period_id", "dim_period", "period_id", "many-to-one"),
    ("fact_deal_credit", "rep_id", "dim_rep", "rep_id", "many-to-one"),
    ("fact_deal_credit", "period_id", "dim_period", "period_id", "many-to-one"),
    ("fact_team_snapshot", "team_id", "dim_team", "team_id", "many-to-one"),
    ("fact_marketing_rep_period", "rep_id", "dim_marketing_rep", "rep_id", "many-to-one"),
    ("fact_marketing_rep_period", "period_id", "dim_period", "period_id", "many-to-one"),
    ("fact_marketing_tour_payout", "rep_id", "dim_marketing_rep", "rep_id", "many-to-one"),
    ("fact_marketing_tour_payout", "period_id", "dim_period", "period_id", "many-to-one"),
    ("fact_marketing_tour_payout", "guest_id", "dim_guest", "guest_id", "many-to-one"),
    ("bridge_tour_guest", "tour_id", "fact_marketing_tour_payout", "tour_id", "many-to-one"),
    ("bridge_tour_guest", "guest_id", "dim_guest", "guest_id", "many-to-one"),
    ("dim_guest", "household_id", "dim_household", "household_id", "many-to-one"),
    ("scenario_result", "scenario_id", "scenario_run", "scenario_id", "one-to-one"),
    ("scenario_payout_series", "scenario_id", "scenario_run", "scenario_id", "many-to-one"),
    ("dim_finance_period", "period_id", "dim_period", "period_id", "one-to-one"),
]

SOURCE_SYSTEMS = [
    ("edw_dev_cognos.cognos_fm", "it_smt_detail", "Tour transactions, show/qualify flags, volume", "Marketing + field tour facts"),
    ("edw_dev_cognos.cognos_fm", "it_smt_marketing", "Tour marketing attributes, office, channel", "Marketing dimensions"),
    ("edw_dev_cognos.cognos_fm", "it_smt_personnel", "Salesperson assignments per tour", "Rep attribution"),
    ("edw_dev_cognos.cognos_fm", "it_smt_contract", "Contract linkage on tours", "Deal / contract spine"),
    ("edw_dev_cognos.cognos_fm", "it_uni_contract", "Unified contract dimension", "Contract status"),
    ("edw_dev_cognos.cognos_fm", "it_uni_lead", "Lead / guest master", "Guest registry views"),
    ("edw_dev_hris.pwcmodels", "commissions", "PwC Varicent commission payments", "Field rep payout, dim_rep, dim_period"),
]


def sheet_overview(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "HGV Compensation Hub — Data Model"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("Catalog", CATALOG),
        ("Schema", SCHEMA),
        ("Full name", f"{CATALOG}.{SCHEMA}"),
        ("Generated from", "data/comp/edw_dev_hris/*.sql DDL"),
        ("Production path", "12_bootstrap_live_source_views.sql → 15 governance → 16 materialize (FY2025–26)"),
        ("App catalog env", "COMP_CATALOG=edw_dev_hris, COMP_SCHEMA=hgv_comp, COMP_DATA_MODE=production"),
        ("Materialized marketing core", "dim_marketing_rep, fact_marketing_tour_payout, fact_marketing_rep_period, dim_period, dim_rep"),
        ("Writable Delta tables", "scenario_*, fact_comp_admin_log, industry_comp_benchmark, plan_assessment_*, fact_regional_bonus_*"),
        ("Data window (script 16)", "2025-01-01 through 2026-12-31"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    auto_width(ws)


def sheet_entities(wb: Workbook) -> None:
    ws = wb.create_sheet("Entities")
    cols = ["Entity", "Object Type", "Layer", "Production Storage", "Grain", "Description", "DDL Source"]
    hdr(ws, 1, cols)
    for r, row in enumerate(ENTITIES, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_columns(wb: Workbook) -> None:
    ws = wb.create_sheet("Columns")
    cols = ["Entity", "Column", "Data Type", "Nullable", "PK", "Description"]
    hdr(ws, 1, cols)
    for r, row in enumerate(COLUMNS, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_relationships(wb: Workbook) -> None:
    ws = wb.create_sheet("Relationships")
    cols = ["From Entity", "From Column", "To Entity", "To Column", "Cardinality"]
    hdr(ws, 1, cols)
    for r, row in enumerate(RELATIONSHIPS, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    auto_width(ws)


def sheet_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Source Systems")
    cols = ["Source Catalog.Schema", "Table", "Description", "Used For"]
    hdr(ws, 1, cols)
    for r, row in enumerate(SOURCE_SYSTEMS, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    auto_width(ws)


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_overview(wb)
    sheet_entities(wb)
    sheet_columns(wb)
    sheet_relationships(wb)
    sheet_sources(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
